import logging
from datetime import datetime
from collections import Counter

from config import (
    DEFAULT_CAGE, DEFAULT_CONDITION, DEFAULT_STORAGE_LOCATION,
    DEFAULT_UOM, SYSTEM_NAME, SYSTEM_SHORT,
)

logger = logging.getLogger(__name__)

# Supabase client — lazy initialization
_supabase = None

def _get_supabase():
    global _supabase
    if _supabase:
        return _supabase
    try:
        from supabase import create_client
        import os
        url = os.environ["SUPABASE_URL"]
        key = os.environ["SUPABASE_SECRET_KEY"]
        _supabase = create_client(url, key)
    except Exception as e:
        logger.error("supabase init failed: %s", e)
        raise
    return _supabase


def _gen_id(n):
    return f"{SYSTEM_SHORT}-{n:05d}"


def _next_id():
    db = _get_supabase()
    res = db.table("scan_log").select("id", count="exact").execute()
    return _gen_id((res.count or 0) + 1)


def append_scan_log(record):
    db  = _get_supabase()
    inv = _next_id()
    qty   = int(record.get("quantity", 0) or 0)
    price = float(record.get("unit_price", 0) or 0)

    row = {
        "inventory_id":    inv,
        "scan_date":       record.get("scan_date", datetime.now().isoformat()),
        "operator":        record.get("operator", ""),
        "user_id":         str(record.get("user_id", "")),
        "nsn":             record.get("nsn", "N/A"),
        "mpn":             record.get("mpn", record.get("part_number", "N/A")),
        "cage_code":       record.get("cage_code", DEFAULT_CAGE),
        "part_name":       record.get("part_name", record.get("name", "")),
        "category_section":record.get("category_section", record.get("category", "")),
        "quantity":        qty,
        "uom":             record.get("uom", record.get("unit", DEFAULT_UOM)),
        "storage_location":record.get("storage_location", DEFAULT_STORAGE_LOCATION),
        "condition":       record.get("condition", DEFAULT_CONDITION),
        "unit_price":      price,
        "total_value":     qty * price,
        "photo_url":       record.get("photo_url", ""),
        "data_source":     record.get("data_source", "checkin"),
    }

    db.table("scan_log").insert(row).execute()
    logger.info("scan_log: inserted %s", inv)
    return inv


def update_inventory(record):
    db    = _get_supabase()
    nsn   = record.get("nsn", "N/A")
    qty   = int(record.get("quantity", 0) or 0)
    price = float(record.get("unit_price", 0) or 0)
    now   = datetime.now().isoformat()
    loc   = record.get("storage_location", DEFAULT_STORAGE_LOCATION)
    photo = record.get("photo_url", "")

    existing = db.table("inventory").select("id,quantity,inventory_id").eq("nsn", nsn).execute()

    if existing.data:
        row = existing.data[0]
        new_qty = row["quantity"] + qty
        db.table("inventory").eq("id", row["id"]).update({
            "quantity":        new_qty,
            "total_value":     new_qty * price if price else 0,
            "storage_location":loc or row.get("storage_location"),
            "unit_price":      price or row.get("unit_price", 0),
            "condition":       record.get("condition", DEFAULT_CONDITION),
            "photo_url":       photo or row.get("photo_url", ""),
            "last_updated":    now,
            "last_operator":   record.get("operator", ""),
        }).execute()
    else:
        count = db.table("inventory").select("id", count="exact").execute()
        inv_id = _gen_id((count.count or 0) + 1)
        db.table("inventory").insert({
            "inventory_id":    inv_id,
            "nsn":             nsn,
            "mpn":             record.get("mpn", record.get("part_number", "N/A")),
            "cage_code":       record.get("cage_code", DEFAULT_CAGE),
            "part_name":       record.get("part_name", record.get("name", "")),
            "category_section":record.get("category_section", record.get("category", "")),
            "quantity":        qty,
            "uom":             record.get("uom", record.get("unit", DEFAULT_UOM)),
            "storage_location":loc,
            "unit_price":      price,
            "total_value":     qty * price,
            "condition":       record.get("condition", DEFAULT_CONDITION),
            "photo_url":       photo,
            "last_updated":    now,
            "last_operator":   record.get("operator", ""),
        }).execute()

    logger.info("inventory: upserted nsn=%s", nsn)


def get_dashboard_summary():
    db = _get_supabase()
    rows = db.table("inventory").select("*").execute().data

    if not rows:
        return f"📦 {SYSTEM_NAME}\n\nInventory is empty. Start with /checkin"

    total_val = sum(float(r.get("total_value") or 0) for r in rows)
    total_qty = sum(int(r.get("quantity") or 0) for r in rows)
    by_cat    = {}
    top       = []

    for r in rows:
        cat = r.get("category_section") or "Other"
        val = float(r.get("total_value") or 0)
        by_cat[cat] = by_cat.get(cat, 0.0) + val
        top.append((r.get("part_name") or "", val))

    top.sort(key=lambda x: x[1], reverse=True)

    lines = [
        f"📊 <b>{SYSTEM_NAME}</b>",
        f"🔢 Positions: <b>{len(rows):,}</b>  |  Units: <b>{total_qty:,}</b>",
        f"💰 <b>Total: ${total_val:,.2f}</b>",
        "",
        "<b>By category:</b>",
    ]
    for cat, val in sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:8]:
        pct = val / total_val * 100 if total_val else 0
        lines.append(f"  • {cat}: ${val:,.2f} ({pct:.1f}%)")

    lines += ["", "<b>Top 5:</b>"]
    for i, (name, val) in enumerate(top[:5], 1):
        lines.append(f"  {i}. {name[:40]}: ${val:,.2f}")

    lines.append(f"\n<i>{datetime.now().strftime('%Y-%m-%d %H:%M')}</i>")
    return "\n".join(lines)


def export_audit_excel(records, output_path):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    now = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    H_FILL = PatternFill("solid", fgColor="1A3A5C")
    H_FONT = Font(color="FFFFFF", bold=True, size=9)
    ALT    = PatternFill("solid", fgColor="EBF5FB")
    TOT    = PatternFill("solid", fgColor="1E8449")
    TOT_F  = Font(bold=True, color="FFFFFF", size=10)
    THIN   = Border(
        left=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"),
    )
    CTR = Alignment(horizontal="center", vertical="center")

    def hdr(ws, row, n):
        for c in range(1, n + 1):
            cell = ws.cell(row, c)
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = CTR
            cell.border = THIN
        ws.row_dimensions[row].height = 22

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Cover Page
    ws_cov = wb.active
    ws_cov.title = "Cover Page"
    ws_cov.column_dimensions["A"].width = 4
    ws_cov.column_dimensions["B"].width = 36
    ws_cov.column_dimensions["C"].width = 40

    ws_cov.merge_cells("B2:C3")
    ws_cov["B2"].value = SYSTEM_NAME.upper()
    ws_cov["B2"].font = Font(bold=True, size=18, color="FFFFFF")
    ws_cov["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cov["B2"].fill = PatternFill("solid", fgColor="1A3A5C")
    ws_cov.row_dimensions[2].height = 32
    ws_cov.row_dimensions[3].height = 32

    ws_cov.merge_cells("B4:C4")
    ws_cov["B4"].value = "PARTS INVENTORY — AUDIT PACKAGE"
    ws_cov["B4"].font = Font(bold=True, size=12, color="7B241C")
    ws_cov["B4"].alignment = Alignment(horizontal="center")
    ws_cov["B4"].fill = PatternFill("solid", fgColor="FADBD8")
    ws_cov.row_dimensions[4].height = 22

    def cov(label, value, bold=False, fill=None):
        r = ws_cov.max_row + 1
        ws_cov.row_dimensions[r].height = 20
        c1 = ws_cov.cell(r, 2, label)
        c2 = ws_cov.cell(r, 3, value)
        c1.font = Font(bold=True, size=10, color="1A3A5C")
        c2.font = Font(bold=bold, size=10)
        if fill:
            c1.fill = PatternFill("solid", fgColor=fill)
            c2.fill = PatternFill("solid", fgColor=fill)

    ws_cov.append([])
    cov("Generated:", now)
    cov("Total positions:", f"{len(records):,}", bold=True, fill="D6E4F0")
    cov("Default condition:", "NOS — New Old Stock")
    cov("ID format:", f"{SYSTEM_SHORT}-XXXXX")
    ws_cov.append([])
    cov("Sheet 2:", f"Inventory Master — {len(records):,} positions")
    cov("Sheet 3:", "Blank Audit Form — print for physical count")
    cov("Sheet 4:", "By Category — summary")
    ws_cov.append([])
    cov("NOTICE:", "READ-ONLY for audit purposes", fill="FEF9E7")

    # Inventory Master
    ws_inv = wb.create_sheet("Inventory Master")
    ws_inv.freeze_panes = "D3"
    ws_inv.sheet_view.zoomScale = 85

    ws_inv.merge_cells("A1:P1")
    ws_inv["A1"].value = f"{SYSTEM_NAME.upper()} — INVENTORY MASTER LIST"
    ws_inv["A1"].font = Font(bold=True, size=13, color="1A3A5C")
    ws_inv["A1"].alignment = Alignment(horizontal="center")
    ws_inv["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws_inv.row_dimensions[1].height = 26

    cols = ["№","Inventory_ID","NSN","MPN","CAGE","Part_Name",
            "Category_Section","Quantity","UOM","Storage_Location",
            "Unit_Price","Total_Value","Condition","Last_Updated","Photo_URL","Remarks"]
    ws_inv.append(cols)
    hdr(ws_inv, 2, len(cols))

    total_val = 0.0
    for i, r in enumerate(records, 1):
        price = float(r.get("unit_price", 0) or 0)
        qty   = int(r.get("quantity", 0) or 0)
        total = price * qty
        total_val += total
        ws_inv.append([
            i,
            r.get("inventory_id", ""),
            r.get("nsn", ""),
            r.get("part_number", r.get("mpn", "N/A")),
            r.get("cage_code", DEFAULT_CAGE),
            r.get("name", r.get("part_name", ""))[:60],
            r.get("category", r.get("category_section", "")),
            qty or "",
            r.get("unit", r.get("uom", DEFAULT_UOM)),
            r.get("storage_location", DEFAULT_STORAGE_LOCATION),
            price or "",
            total or "",
            r.get("condition", DEFAULT_CONDITION),
            r.get("last_updated", ""),
            r.get("photo_url", ""),
            "",
        ])
        rn = i + 2
        if i % 2 == 0:
            for c in range(1, len(cols) + 1):
                ws_inv.cell(rn, c).fill = ALT
        for c in range(1, len(cols) + 1):
            ws_inv.cell(rn, c).border = THIN
        ws_inv.row_dimensions[rn].height = 15

    tr = len(records) + 3
    ws_inv.cell(tr, 1, "TOTAL").font = TOT_F
    ws_inv.cell(tr, 6, f"{len(records):,} positions").font = TOT_F
    ws_inv.cell(tr, 12, total_val or "N/A").font = TOT_F
    for c in range(1, len(cols) + 1):
        ws_inv.cell(tr, c).fill = TOT
    ws_inv.row_dimensions[tr].height = 20
    set_widths(ws_inv, [5,12,19,16,6,50,22,8,6,14,10,12,12,14,30,16])

    # Blank Audit Form
    ws_aud = wb.create_sheet("Blank Audit Form")
    ws_aud.freeze_panes = "D3"
    ws_aud.merge_cells("A1:K1")
    ws_aud["A1"].value = "PHYSICAL COUNT SHEET — fill in during physical audit"
    ws_aud["A1"].font = Font(bold=True, size=12, color="7B241C")
    ws_aud["A1"].alignment = Alignment(horizontal="center")
    ws_aud["A1"].fill = PatternFill("solid", fgColor="FADBD8")
    ws_aud.row_dimensions[1].height = 24

    acols = ["№","Inventory_ID","NSN","MPN","Part_Name",
             "Storage_Location","Physical_Count","UOM","Condition_Found","Match (Y/N)","Notes"]
    ws_aud.append(acols)
    hdr(ws_aud, 2, len(acols))
    for i, r in enumerate(records, 1):
        ws_aud.append([i, r.get("inventory_id",""), r.get("nsn",""),
                       r.get("part_number", r.get("mpn","N/A")),
                       r.get("name","")[:55], "","", r.get("unit", DEFAULT_UOM),"","",""])
        rn = i + 2
        if i % 2 == 0:
            for c in range(1, len(acols)+1):
                ws_aud.cell(rn, c).fill = ALT
        for c in range(1, len(acols)+1):
            ws_aud.cell(rn, c).border = THIN
        ws_aud.row_dimensions[rn].height = 15
    set_widths(ws_aud, [5,12,19,16,52,16,14,6,15,10,20])

    # By Category
    ws_cat = wb.create_sheet("By Category")
    ws_cat.merge_cells("A1:D1")
    ws_cat["A1"].value = "INVENTORY SUMMARY BY CATEGORY"
    ws_cat["A1"].font = Font(bold=True, size=13, color="1A3A5C")
    ws_cat["A1"].alignment = Alignment(horizontal="center")
    ws_cat["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
    ws_cat.row_dimensions[1].height = 26
    ws_cat.append(["Category_Section","Part_Count","Total_Value ($)","Notes"])
    hdr(ws_cat, 2, 4)
    set_widths(ws_cat, [40,12,18,30])

    cat_counts = Counter(r.get("category", r.get("category_section","Other")) for r in records)
    cat_vals   = {}
    for r in records:
        cat = r.get("category", r.get("category_section","Other"))
        cat_vals[cat] = cat_vals.get(cat,0.0) + float(r.get("unit_price",0) or 0) * int(r.get("quantity",0) or 0)

    for i, (cat, cnt) in enumerate(cat_counts.most_common(), 1):
        ws_cat.append([cat, cnt, cat_vals.get(cat,0) or "N/A",""])
        rn = i + 2
        if i % 2 == 0:
            for c in range(1,5): ws_cat.cell(rn,c).fill = ALT
        for c in range(1,5): ws_cat.cell(rn,c).border = THIN

    tr2 = ws_cat.max_row + 2
    ws_cat.cell(tr2,1,"TOTAL").font = TOT_F
    ws_cat.cell(tr2,2,len(records)).font = TOT_F
    ws_cat.cell(tr2,3,sum(cat_vals.values()) or "N/A").font = TOT_F
    for c in range(1,5): ws_cat.cell(tr2,c).fill = TOT

    wb.save(output_path)
    return output_path
